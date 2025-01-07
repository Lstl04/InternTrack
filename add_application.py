import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

application_file = r"C:\Users\louis\Documents\Intership_application.xlsx"

def add_application():
    company = input("What company did you apply to: ")
    job = input("What is the job title: ")
    date = datetime.today().strftime('%d/%m/%Y')
    link = input("Link to application (press return if none): ")
    loc = input("Enter the period of the internship and the location in the format 'period - location': ")
    
    df = pd.read_excel(application_file)
    
    new_row = pd.DataFrame([{
        "Company": company,
        "Poste": job,
        "date applied": date,
        "link to application": link if link else "N/A",  
        "intership date": loc,
        "Progress": ""  
    }])
    
    df = pd.concat([df, new_row], ignore_index=True)
    df.to_excel(application_file, index=False, engine='openpyxl')
    
    wb = load_workbook(application_file)
    sheet = wb.active

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    new_row_index = len(df)

    progress_column = df.columns.get_loc("Progress") + 1
    sheet.cell(row=new_row_index + 1, column=progress_column).fill = yellow_fill

    wb.save(application_file)

    print("Application information added successfully")


