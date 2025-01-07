import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

received_mails = pd.read_csv("mails.csv")
applications = r"C:\Users\louis\Documents\Intership_application.xlsx"
filtered_rows = received_mails[received_mails["Internship?"] == "Yes"]
emails_data = filtered_rows[["Company name", "Application progress"]]

wb = load_workbook(applications)
sheet = wb.active

blue_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")  # Blue
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red

for _, email_row in emails_data.iterrows():
    company_name = email_row["Company name"]
    application_progress = email_row["Application progress"]

    for excel_row in range(2, sheet.max_row + 1):  
        excel_company = sheet.cell(row=excel_row, column=1).value  
        if excel_company == company_name:
            progress_column = 6 
            if application_progress == 2:
                sheet.cell(row=excel_row, column=progress_column).fill = blue_fill
            elif application_progress == 3:
                sheet.cell(row=excel_row, column=progress_column).fill = red_fill
            break  

wb.save(applications)

print("Applications updated successfully!")
